"""
Generate an Excel preview of invoices to be created.

Usage:
    python3 preview_invoices_excel.py transaction-history.csv -o invoices_preview.xlsx
"""

import sys
from typing import List, Dict

import click
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from invoice_templates import InvoiceGenerator


def parse_wise_csv(csv_path: str) -> List[Dict]:
    """Parse Wise transaction CSV and extract incoming payments."""
    import csv

    transactions = []

    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)

        for row in reader:
            direction = row.get("Direction", "").strip()
            status = row.get("Status", "").strip()
            source_name = row.get("Source name", "").strip()

            # Only process completed incoming transfers
            if direction != "IN" or status != "COMPLETED":
                continue

            # Only process known clients
            if "GILAD WEINBERG" not in source_name.upper() and \
               "AMZEXPERTGLOBALL" not in source_name.upper():
                continue

            # Extract relevant data
            created_date = row.get("Created on", "").strip()
            amount_str = row.get("Source amount (after fees)", "").strip()

            # Parse date (format: "2026-03-06 19:57:01")
            invoice_date = created_date.split(" ")[0] if created_date else ""

            # Parse amount
            try:
                amount = float(amount_str) if amount_str else 0.0
            except ValueError:
                continue

            transactions.append({
                "client_name": source_name,
                "date": invoice_date,
                "amount": amount,
                "reference": row.get("Reference", "").strip(),
            })

    return transactions


@click.command()
@click.argument("csv_file", type=click.Path(exists=True))
@click.option("-o", "--output", default="invoices_preview.xlsx", help="Output Excel file")
def main(csv_file: str, output: str):
    """Generate Excel preview of invoices to be created."""

    # Parse transactions
    click.echo(f"Parsing {csv_file}...")
    transactions = parse_wise_csv(csv_file)

    if not transactions:
        click.echo("No incoming transactions found from known clients.")
        return

    click.echo(f"Found {len(transactions)} incoming transactions.\n")

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice Preview"

    # Define styles
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    currency_font = Font(size=10)
    text_font = Font(size=10)

    gilad_fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
    amz_fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    headers = [
        "Invoice #",
        "Client",
        "Date",
        "Wire Amount",
        "Line 1 - Description",
        "Line 1 - Amount",
        "Line 2 - Description",
        "Line 2 - Amount",
        "Total Invoice",
        "Notes"
    ]

    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Set column widths
    ws.column_dimensions['A'].width = 15  # Invoice #
    ws.column_dimensions['B'].width = 20  # Client
    ws.column_dimensions['C'].width = 12  # Date
    ws.column_dimensions['D'].width = 14  # Wire Amount
    ws.column_dimensions['E'].width = 35  # Line 1 Desc
    ws.column_dimensions['F'].width = 14  # Line 1 Amount
    ws.column_dimensions['G'].width = 30  # Line 2 Desc
    ws.column_dimensions['H'].width = 14  # Line 2 Amount
    ws.column_dimensions['I'].width = 14  # Total
    ws.column_dimensions['J'].width = 50  # Notes

    # Process transactions
    row_num = 2
    total_gilad = 0.0
    total_amz = 0.0

    for txn in transactions:
        try:
            # Generate invoice data
            invoice_data = InvoiceGenerator.generate_invoice_data(
                client_name=txn["client_name"],
                wire_amount=txn["amount"],
                wire_date=txn["date"],
                customer_id="<placeholder>"
            )

            # Extract line items
            line_items = invoice_data["line_items"]

            # Determine fill color
            is_gilad = "GILAD" in txn["client_name"].upper()
            fill = gilad_fill if is_gilad else amz_fill

            if is_gilad:
                total_gilad += txn["amount"]
            else:
                total_amz += txn["amount"]

            # Write data
            ws.cell(row=row_num, column=1, value=invoice_data["invoice_number"])
            ws.cell(row=row_num, column=2, value=txn["client_name"])
            ws.cell(row=row_num, column=3, value=txn["date"])
            ws.cell(row=row_num, column=4, value=txn["amount"])

            # Line 1 - Include name + description (Ref line)
            line1_desc = line_items[0]["name"]
            if line_items[0].get("description"):
                line1_desc += f"\n{line_items[0]['description']}"
            ws.cell(row=row_num, column=5, value=line1_desc)
            ws.cell(row=row_num, column=6, value=line_items[0]["rate"])

            if len(line_items) == 2:
                # Line 2 - Include name + description (Ref line)
                line2_desc = line_items[1]["name"]
                if line_items[1].get("description"):
                    line2_desc += f"\n{line_items[1]['description']}"
                ws.cell(row=row_num, column=7, value=line2_desc)
                ws.cell(row=row_num, column=8, value=line_items[1]["rate"])
            else:
                ws.cell(row=row_num, column=7, value="")
                ws.cell(row=row_num, column=8, value="")

            ws.cell(row=row_num, column=9, value=txn["amount"])
            ws.cell(row=row_num, column=10, value=invoice_data["notes"])

            # Apply formatting
            for col_num in range(1, 11):
                cell = ws.cell(row=row_num, column=col_num)
                cell.fill = fill
                cell.border = thin_border
                cell.font = text_font

                # Currency formatting
                if col_num in [4, 6, 8, 9]:
                    cell.number_format = '$#,##0.00'
                    cell.alignment = Alignment(horizontal="right")
                elif col_num == 3:  # Date
                    cell.alignment = Alignment(horizontal="center")
                else:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

            row_num += 1

        except Exception as e:
            click.echo(f"Error processing {txn['client_name']} on {txn['date']}: {e}")
            continue

    # Add summary rows
    row_num += 1
    summary_font = Font(bold=True, size=11)
    summary_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    # Gilad total
    ws.cell(row=row_num, column=1, value="Gilad Weinberg Total:")
    ws.cell(row=row_num, column=1).font = summary_font
    ws.cell(row=row_num, column=4, value=total_gilad)
    ws.cell(row=row_num, column=4).font = summary_font
    ws.cell(row=row_num, column=4).number_format = '$#,##0.00'
    ws.cell(row=row_num, column=4).fill = summary_fill

    row_num += 1

    # Amz total
    ws.cell(row=row_num, column=1, value="Amz-expert Global Total:")
    ws.cell(row=row_num, column=1).font = summary_font
    ws.cell(row=row_num, column=4, value=total_amz)
    ws.cell(row=row_num, column=4).font = summary_font
    ws.cell(row=row_num, column=4).number_format = '$#,##0.00'
    ws.cell(row=row_num, column=4).fill = summary_fill

    row_num += 1

    # Grand total
    grand_font = Font(bold=True, size=12, color="FFFFFF")
    grand_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    ws.cell(row=row_num, column=1, value="GRAND TOTAL:")
    ws.cell(row=row_num, column=1).font = grand_font
    ws.cell(row=row_num, column=1).fill = grand_fill
    ws.cell(row=row_num, column=4, value=total_gilad + total_amz)
    ws.cell(row=row_num, column=4).font = grand_font
    ws.cell(row=row_num, column=4).fill = grand_fill
    ws.cell(row=row_num, column=4).number_format = '$#,##0.00'

    # Freeze header row
    ws.freeze_panes = "A2"

    # Save workbook
    wb.save(output)

    click.echo(f"\n✓ Excel preview saved to: {output}")
    click.echo(f"  Total invoices: {len(transactions)}")
    click.echo(f"  Gilad Weinberg: ${total_gilad:,.2f}")
    click.echo(f"  Amz-expert Global: ${total_amz:,.2f}")
    click.echo(f"  Grand Total: ${total_gilad + total_amz:,.2f}")


if __name__ == "__main__":
    main()
